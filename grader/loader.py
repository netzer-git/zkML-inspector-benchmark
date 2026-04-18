"""Load ground truth (xlsx) and agent output (JSON), parse and validate."""

from __future__ import annotations

import json
import re
from collections import defaultdict
from dataclasses import dataclass, field
from pathlib import Path
from typing import NamedTuple

import openpyxl

from grader import CATEGORIES, SECURITY_CONCERNS, SEVERITIES


class CodeRef(NamedTuple):
    """A reference to a code location: file, optional start/end lines."""
    filename: str
    start_line: int | None = None
    end_line: int | None = None


@dataclass
class GroundTruthFinding:
    entry_id: str
    issue_id: str
    issue_name: str
    issue_explanation: str
    severity: str
    category: str
    security_concern: str
    relevant_code: list[CodeRef] = field(default_factory=list)
    paper_reference: str = ""


@dataclass
class AgentFinding:
    entry_id: str
    issue_name: str
    issue_explanation: str
    severity: str
    category: str
    security_concern: str
    relevant_code: list[CodeRef] = field(default_factory=list)
    paper_reference: str = ""


# ---------------------------------------------------------------------------
# Parsing helpers
# ---------------------------------------------------------------------------

_CODE_REF_PATTERN = re.compile(
    r"([A-Za-z0-9_.\/\-]+\.[A-Za-z0-9]+)"  # filename with extension
    r"(?::(\d+)(?:\s*[-–]\s*(\d+))?)?"       # optional :start[-end]
)


def parse_code_refs(raw: str | None) -> list[CodeRef]:
    """Parse a code-reference string like 'file.rs:10-15, other.cu:3' into CodeRefs."""
    if not raw or raw.strip().lower() in ("none", "-", ""):
        return []
    refs: list[CodeRef] = []
    # Split on comma, semicolon, or standalone whitespace between refs
    for segment in re.split(r"[,;]\s*", raw.strip()):
        segment = segment.strip()
        if not segment:
            continue
        m = _CODE_REF_PATTERN.search(segment)
        if m:
            filename = m.group(1)
            start = int(m.group(2)) if m.group(2) else None
            end = int(m.group(3)) if m.group(3) else start
            refs.append(CodeRef(filename, start, end))
    return refs


def _normalize_entry_id(raw: str) -> str:
    """Normalize project name to lowercase for consistent keying."""
    return raw.strip().lower()


def _validate_severity(value: str, context: str) -> str:
    if value not in SEVERITIES:
        raise ValueError(f"{context}: invalid severity '{value}'. Must be one of {SEVERITIES}")
    return value


def _validate_category(value: str, context: str) -> str:
    if value not in CATEGORIES:
        raise ValueError(f"{context}: invalid category '{value}'. Must be one of {CATEGORIES}")
    return value


def _validate_security_concern(value: str, context: str) -> str:
    if value not in SECURITY_CONCERNS:
        raise ValueError(
            f"{context}: invalid security-concern '{value}'. Must be one of {SECURITY_CONCERNS}"
        )
    return value


# ---------------------------------------------------------------------------
# Ground truth loader (xlsx)
# ---------------------------------------------------------------------------

_EXPECTED_HEADERS = [
    "entry-id", "issue-id", "issue-name", "issue-explanation",
    "severity", "category", "security-concern", "relevant-code", "paper-reference",
]


def load_ground_truth(xlsx_path: str | Path) -> dict[str, list[GroundTruthFinding]]:
    """Load ground truth from xlsx. Returns findings grouped by normalized entry_id."""
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip().lower() for cell in next(ws.iter_rows(max_row=1))]
    for expected in _EXPECTED_HEADERS:
        if expected not in headers:
            raise ValueError(f"Missing expected column '{expected}' in xlsx. Found: {headers}")

    col_idx = {h: i for i, h in enumerate(headers)}
    result: dict[str, list[GroundTruthFinding]] = defaultdict(list)

    for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        entry_id_raw = row[col_idx["entry-id"]]
        if entry_id_raw is None:
            continue

        entry_id = str(entry_id_raw).strip()
        issue_id = str(row[col_idx["issue-id"]] or "").strip()
        context = f"Row {row_num} ({issue_id or entry_id})"

        severity_raw = str(row[col_idx["severity"]] or "").strip()
        category_raw = str(row[col_idx["category"]] or "").strip()
        concern_raw = str(row[col_idx["security-concern"]] or "").strip()

        finding = GroundTruthFinding(
            entry_id=entry_id,
            issue_id=issue_id,
            issue_name=str(row[col_idx["issue-name"]] or "").strip(),
            issue_explanation=str(row[col_idx["issue-explanation"]] or "").strip(),
            severity=_validate_severity(severity_raw, context),
            category=_validate_category(category_raw, context),
            security_concern=_validate_security_concern(concern_raw, context),
            relevant_code=parse_code_refs(str(row[col_idx["relevant-code"]] or "")),
            paper_reference=str(row[col_idx["paper-reference"]] or "").strip(),
        )
        result[_normalize_entry_id(entry_id)].append(finding)

    wb.close()
    return dict(result)


# ---------------------------------------------------------------------------
# Agent output loader (JSON)
# ---------------------------------------------------------------------------

_REQUIRED_AGENT_FIELDS = {
    "entry-id", "issue-name", "issue-explanation",
    "severity", "category", "security-concern",
    "relevant-code", "paper-reference",
}


def load_agent_output(json_path: str | Path) -> dict[str, list[AgentFinding]]:
    """Load agent output from a flat JSON array. Returns findings grouped by normalized entry_id."""
    with open(json_path, encoding="utf-8") as f:
        data = json.load(f)

    if not isinstance(data, list):
        raise ValueError("Agent output must be a JSON array of finding objects")

    result: dict[str, list[AgentFinding]] = defaultdict(list)

    for i, obj in enumerate(data):
        if not isinstance(obj, dict):
            raise ValueError(f"Agent finding #{i}: expected object, got {type(obj).__name__}")

        missing = _REQUIRED_AGENT_FIELDS - set(obj.keys())
        if missing:
            raise ValueError(f"Agent finding #{i}: missing required fields: {missing}")

        entry_id_raw = str(obj["entry-id"]).strip()
        context = f"Agent finding #{i} ({entry_id_raw})"

        severity = str(obj["severity"]).strip()
        category = str(obj["category"]).strip()
        concern = str(obj["security-concern"]).strip()

        finding = AgentFinding(
            entry_id=entry_id_raw,
            issue_name=str(obj["issue-name"]).strip(),
            issue_explanation=str(obj["issue-explanation"]).strip(),
            severity=_validate_severity(severity, context),
            category=_validate_category(category, context),
            security_concern=_validate_security_concern(concern, context),
            relevant_code=parse_code_refs(str(obj.get("relevant-code", "") or "")),
            paper_reference=str(obj.get("paper-reference", "") or "").strip(),
        )
        result[_normalize_entry_id(entry_id_raw)].append(finding)

    return dict(result)
